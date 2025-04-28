import pandas as pd
import os
import time
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel, from_excel
from datetime import datetime
from tqdm import tqdm

start_time = time.time()    #Tracking process time

# Script Header Information and Description
print("=" * 40)
print("\U0001F4E6 PowerBI Sales Cube Updater")
print("\U0001F4E6 PowerBI Sales Cube Aktualisierer")
print("=" * 40)
print("Version: 1.4.7")
print("Letzte Aktualisierung / Last updated: April 2025")
print("Autor / Author: George Schwenzfeger")
print("Kontakt / Contact: gschwenzfeger@bryanmedical.net\n")

# English
print("\U0001F4D8 Description:")
print("This script imports sales data from a selected CSV file and appends it to an existing Excel file.")
print("\u2705 Skips rows with dates already present in the Excel.")
print("\u2705 Matches and skips header row if already in Excel.")
print("\u2705 Provides a progress bar while updating.")
print("\u2705 Logs every update with bilingual summary.\n")

# German
print("\U0001F4D7 Beschreibung:")
print("Dieses Skript importiert Verkaufsdaten aus einer ausgewählten CSV-Datei und fügt sie einer vorhandenen Excel-Datei hinzu.")
print("\u2705 Überspringt Zeilen mit Datumswerten, die bereits in der Excel-Datei vorhanden sind.")
print("\u2705 Erkennt die Kopfzeile und überspringt sie bei Übereinstimmung.")
print("\u2705 Zeigt eine Fortschrittsanzeige während des Updates.")
print("\u2705 Protokolliert jeden Import mit zweisprachiger Zusammenfassung.\n")

try:
    # Bring up Windows file chooser:
    root = Tk()
    root.withdraw()

    # Choose which CSV file to upload:
    csv_path = filedialog.askopenfilename(
        title="Select CSV File / Wähle die CSV-Datei aus",
        filetypes=[("CSV files", "*.csv")]
    )
    if not csv_path:
        print("\u274C No CSV file selected.\n\u274C Keine CSV-Datei ausgewählt.")
        input("\U0001F51A Press ENTER to exit / Drücke ENTER zum Beenden...")
        raise SystemExit()

    print("\n\U0001F4C4 CSV file selected:\n" + csv_path)
    print("\U0001F4C4 CSV-Datei ausgewählt:\n" + csv_path)

    # Choose which Excel file to append the CSV file data to:
    excel_path = filedialog.askopenfilename(
        title="Select Excel File / Wähle die Excel-Datei aus",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not excel_path:
        print("\u274C No Excel file selected.\n\u274C Keine Excel-Datei ausgewählt.")
        input("\U0001F51A Press ENTER to exit / Drücke ENTER zum Beenden...")
        raise SystemExit()

    
    print("\n\U0001F4C8 Excel file selected:\n" + excel_path)
    print("\U0001F4C8 Excel-Datei ausgewählt:\n" + excel_path + "\n")

    # Load CSV data into a dataframe
    csv_df = pd.read_csv(csv_path)
    csv_headers = list(csv_df.columns)

    # Load Excel file for manipulation
    wb = load_workbook(excel_path)
    sheet = wb.active

    # Find the true next empty row in the Excel file to avoid empty rows
    for row in range(sheet.max_row, 0, -1):
        cell_val = sheet.cell(row=row, column=1).value
        if cell_val is not None and str(cell_val).strip() != "":
            start_row = row + 1
            break
    else:
        start_row = 2

    excel_headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Normalize both file headers to compare and make sure they match
    def normalize_headers(header_row):
        return [str(h).strip().lower() for h in header_row]

    norm_csv_headers = normalize_headers(csv_headers)
    norm_excel_headers = normalize_headers(excel_headers)

    if norm_csv_headers == norm_excel_headers:
        print("\U0001F9E0 Headers match. CSV header row will be skipped.")
        print("\U0001F9E0 Überschriften stimmen überein. Kopfzeile wird übersprungen.\n")
        csv_data = csv_df.iloc[1:-1].copy()  # skip header and total line
        header_log = "\U0001F9E0 Headers matched – CSV header row skipped.\n\U0001F9E0 Überschriften stimmen überein – CSV-Kopfzeile übersprungen.\n"
    else:
        print("\u26A0️ Header mismatch detected. Proceeding anyway.")
        print("\u26A0️ Überschriften stimmen nicht überein. Fortfahren...\n")
        print("CSV Headers:\n" + ", ".join(csv_headers))
        print("Excel Headers:\n" + ", ".join(str(h) for h in excel_headers) + "\n")
        csv_data = csv_df[:-1].copy()  # skip total line only
        header_log = "\u26A0️ Header mismatch – data appended anyway. Please review structure.\n\u26A0️ Überschriften stimmen nicht überein – Daten trotzdem angehängt. Bitte Struktur überprüfen.\n"

    # Create a dataframe from the Excel data
    excel_data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    excel_df = pd.DataFrame(excel_data, columns=excel_headers)

    # Function to force the excel date numeric style for easy conversion
    def parse_excel_date(val):
        try:
            if isinstance(val, (int, float)):
                return from_excel(val).date()
            return pd.to_datetime(val, errors='coerce').date()
        except:
            return None

    # Forcing the numeric excel date
    excel_df['__date_only__'] = excel_df['Date'].apply(parse_excel_date)
    excel_df = excel_df[excel_df['__date_only__'].notnull()]

    csv_data['__date_only__'] = pd.to_datetime(csv_data['Date'], format='mixed', errors='coerce').dt.date
    csv_data = csv_data[csv_data['__date_only__'].notnull()]

    # Checking if the dates of the data being uploaded exist already
    existing_dates = set(excel_df['__date_only__'].unique())
    csv_data['__is_duplicate_date__'] = csv_data['__date_only__'].isin(existing_dates)

    # Keeping track of the dates to skip, and then removing those rows from the upload
    skipped_dates = csv_data.loc[csv_data['__is_duplicate_date__'], '__date_only__'].unique()
    new_data = csv_data.loc[~csv_data['__is_duplicate_date__']].drop(columns=['__date_only__', '__is_duplicate_date__'])

    added_rows = len(new_data)
    skipped_dates_list = [str(d) for d in sorted(skipped_dates)]

    if added_rows == 0:
        print("\U0001F501 All dates in the CSV already exist in the Excel file.")
        print("\U0001F501 Alle Datumswerte aus der CSV-Datei sind bereits vorhanden.\n")
        input("\U0001F51A Press ENTER to exit / Drücke ENTER zum Beenden...")
        raise SystemExit()

    print(f"\U0001F4E5 Appending {added_rows} new rows to Excel...\n\U0001F4E5 Füge {added_rows} neue Zeilen in Excel ein...\n")

    # Appending the new rows of data to the excel file:
    for row_idx, row in tqdm(new_data.iterrows(), total=len(new_data), desc="Updating Excel", unit="row"):
        for col_idx, col_name in enumerate(new_data.columns, start=1):
            value = row[col_name]
            cell = sheet.cell(row=start_row + row_idx - 1, column=col_idx)

            if col_name == 'Date':
                try:
                    parsed_date = pd.to_datetime(value, errors='coerce')
                    if pd.notnull(parsed_date):
                        cell.value = to_excel(parsed_date.to_pydatetime())
                        cell.number_format = '0'
                    else:
                        cell.value = value
                except Exception:
                    cell.value = value
            else:
                cell.value = value

    wb.save(excel_path)
    wb.close()

    # Script performance tracking (optional):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elapsed = time.time() - start_time
    log_message = (
        f"{timestamp} | \u2705 {added_rows} rows added to '{os.path.basename(excel_path)}'. "
        f"⏱️ Duration: {elapsed:.2f} seconds\n"
        f"{timestamp} | \u2705 {added_rows} Zeilen zu '{os.path.basename(excel_path)}' hinzugefügt. "
        f"⏱️ Dauer: {elapsed:.2f} Sekunden\n"
        f"{header_log}"
    )

    if skipped_dates_list:
        skipped_str = ", ".join(skipped_dates_list)
        log_message += (
            f"\u274C Skipped dates (already in Excel): {skipped_str}\n"
            f"\u274C Übersprungene Datumswerte (bereits vorhanden): {skipped_str}\n"
        )

    # Logging what actions were taken
    log_path = os.path.join(os.path.dirname(__file__), "import_log.txt")
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(log_message)

    print("\u2705 Import complete.")
    print("\u2705 Import abgeschlossen.\n")
    print(log_message)

# Exception handling:
except Exception as e:
    import traceback
    print("\n\u274C Error occurred / Fehler ist aufgetreten:")
    traceback.print_exc()

# Holds terminal window open until user confirmation:
input("\U0001F51A Press ENTER to exit / Drücke ENTER zum Beenden...")
