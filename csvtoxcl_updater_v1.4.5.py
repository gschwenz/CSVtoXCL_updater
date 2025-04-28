import pandas as pd
import os
import time
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from datetime import datetime
from tqdm import tqdm

def parse_excel_date(val):
    try:
        # Convert Excel float date → datetime.date
        if isinstance(val, (int, float)):
            return from_excel(val).date()
        return pd.to_datetime(val, errors='coerce').date()
    except:
        return pd.NaT

start_time = time.time()

# Header data
print("=" * 40)
print("📦 PowerBI Sales Cube Updater")
print("📦 PowerBI Sales Cube Aktualisierer")
print("=" * 40)
print("Version: 1.4.5")
print("Letzte Aktualisierung / Last updated: April 2025")
print("Autor / Author: George Schwenzfeger")
print("Kontakt / Contact: gschwenzfeger@bryanmedical.net\n")

# English section
print("📘 Description:")
print("This script imports sales data from a selected CSV file and appends it to an existing Excel file.")
print("✅ Skips rows with dates already present in the Excel.")
print("✅ Matches and skips header row if already in Excel.")
print("✅ Provides a progress bar while updating.")
print("✅ Logs every update with bilingual summary.\n")

# German section
print("📗 Beschreibung:")
print("Dieses Skript importiert Verkaufsdaten aus einer ausgewählten CSV-Datei und fügt sie einer vorhandenen Excel-Datei hinzu.")
print("✅ Überspringt Zeilen mit Datumswerten, die bereits in der Excel-Datei vorhanden sind.")
print("✅ Erkennt die Kopfzeile und überspringt sie bei Übereinstimmung.")
print("✅ Zeigt eine Fortschrittsanzeige während des Updates.")
print("✅ Protokolliert jeden Import mit zweisprachiger Zusammenfassung.\n")


try:
    root = Tk()
    root.withdraw()

    # --- FILE PICKING ---
    csv_path = filedialog.askopenfilename(
        title="Select CSV File / Wähle die CSV-Datei aus",
        filetypes=[("CSV files", "*.csv")]
    )
    if not csv_path:
        print("❌ No CSV file selected.\n❌ Keine CSV-Datei ausgewählt.")
        input("🔚 Press ENTER to exit / Drücke ENTER zum Beenden...")
        raise SystemExit()

    excel_path = filedialog.askopenfilename(
        title="Select Excel File / Wähle die Excel-Datei aus",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not excel_path:
        print("❌ No Excel file selected.\n❌ Keine Excel-Datei ausgewählt.")
        input("🔚 Press ENTER to exit / Drücke ENTER zum Beenden...")
        raise SystemExit()

    # --- FILE PATH OUTPUT ---
    print("\n📄 CSV file selected:\n" + csv_path)
    print("📊 Excel file selected:\n" + excel_path)
    print("\n📄 CSV-Datei ausgewählt:\n" + csv_path)
    print("📊 Excel-Datei ausgewählt:\n" + excel_path + "\n")

    # --- LOAD DATA ---
    csv_df = pd.read_csv(csv_path)
    csv_headers = list(csv_df.columns)

    wb = load_workbook(excel_path)
    sheet = wb.active

    # TRUE LAST NON-EMPTY ROW (prevent blank row)
    for row in range(sheet.max_row, 0, -1):
        cell_val = sheet.cell(row=row, column=1).value
        if cell_val is not None and str(cell_val).strip() != "":
            start_row = row + 1
            break
    else:
        start_row = 2

    excel_headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # --- HEADER CHECK ---
    def normalize_headers(header_row):
        return [str(h).strip().lower() for h in header_row]

    norm_csv_headers = normalize_headers(csv_headers)
    norm_excel_headers = normalize_headers(excel_headers)

    csv_data = csv_df.copy()
    if norm_csv_headers == norm_excel_headers:
        print("🧠 Headers match. CSV header row will be skipped.")
        print("🧠 Überschriften stimmen überein. Kopfzeile wird übersprungen.\n")
        header_log = "🧠 Headers matched – CSV header row skipped.\n"
        csv_data = csv_df.iloc[1:].copy()
        csv_data = csv_data[:-1].copy()  # Drop the last row (totaling line)

    else:
        print("⚠️ Header mismatch detected. Proceeding anyway.")
        print("⚠️ Überschriften stimmen nicht überein. Fortfahren...\n")
        print("CSV Headers:\n" + ", ".join(csv_headers))
        print("Excel Headers:\n" + ", ".join(str(h) for h in excel_headers) + "\n")
        header_log = "⚠️ Header mismatch – data appended anyway. Please review structure.\n"

    # --- LOAD EXISTING EXCEL DATA FOR DATE FILTERING ---
    excel_data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    excel_df = pd.DataFrame(excel_data, columns=excel_headers)

    # --- DATE FILTERING ---
    if 'Date' not in csv_data.columns or 'Date' not in excel_df.columns:
        raise ValueError("❌ 'Date' column not found in one of the files.\n❌ Spalte 'Date' fehlt in einer Datei.")

    csv_data['__date_only__'] = pd.to_datetime(csv_data['Date'], format='mixed', errors='coerce').dt.date
    excel_df['__date_only__'] = excel_df['Date'].apply(parse_excel_date)

    if csv_data['__date_only__'].isna().any():
        bad_rows = csv_data[csv_data['__date_only__'].isna()]
        print(f"⚠️ {len(bad_rows)} rows in the CSV have unrecognized date formats and were skipped.")
        print(f"⚠️ {len(bad_rows)} Zeilen in der CSV-Datei enthalten ungültige Datumswerte und wurden übersprungen.\n")
        csv_data = csv_data.dropna(subset=['__date_only__'])

    existing_dates = set(excel_df['__date_only__'].dropna().unique())
    csv_data['__is_duplicate_date__'] = csv_data['__date_only__'].isin(existing_dates)

    skipped_dates = csv_data.loc[csv_data['__is_duplicate_date__'], '__date_only__'].unique()
    new_data = csv_data.loc[~csv_data['__is_duplicate_date__']].drop(columns=['__date_only__', '__is_duplicate_date__'])

    added_rows = len(new_data)
    skipped_dates_list = [str(d) for d in sorted(skipped_dates)]

    if added_rows == 0:
        print("🔁 All dates in the CSV already exist in the Excel file.")
        print("🔁 Alle Datumswerte aus der CSV-Datei sind bereits vorhanden.\n")
        input("🔚 Press ENTER to exit / Drücke ENTER zum Beenden...")
        raise SystemExit()

    # --- WRITE TO EXCEL WITH PROGRESS BAR ---
    print(f"📥 Appending {added_rows} new rows to Excel...\n📥 Füge {added_rows} neue Zeilen in Excel ein...\n")
    for row_idx, row in tqdm(new_data.iterrows(), total=len(new_data), desc="Updating Excel", unit="row"):
        for col_idx, col_name in enumerate(new_data.columns, start=1):
            value = row[col_name]
            cell = sheet.cell(row=start_row + row_idx - 1, column=col_idx)

            if col_name == 'Date':
                try:
                    parsed_date = pd.to_datetime(value, errors='coerce')
                    if pd.notnull(parsed_date):
                        cell.value = to_excel(parsed_date.to_pydatetime())
                        cell.number_format = '0'  # Display as number
                    else:
                        cell.value = value
                except Exception:
                    cell.value = value
            else:
                cell.value = value


    wb.save(excel_path)
    wb.close()

    # --- FINAL LOGGING & SUMMARY ---
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elapsed = time.time() - start_time
    log_message = (
        f"{timestamp} | ✅ {added_rows} rows added to '{os.path.basename(excel_path)}'. "
        f"⏱️ Duration: {elapsed:.2f} seconds\n"
        f"{timestamp} | ✅ {added_rows} Zeilen zu '{os.path.basename(excel_path)}' hinzugefügt. "
        f"⏱️ Dauer: {elapsed:.2f} Sekunden\n"
        f"{header_log}"
    )

    if skipped_dates_list:
        skipped_str = ", ".join(skipped_dates_list)
        log_message += (
            f"❌ Skipped dates (already in Excel): {skipped_str}\n"
            f"❌ Übersprungene Datumswerte (bereits vorhanden): {skipped_str}\n"
        )

    log_path = os.path.join(os.path.dirname(__file__), "import_log.txt")
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(log_message)

    # --- PRINT FINAL SUMMARY TO CONSOLE ---
    print("✅ Import complete.")
    print("✅ Import abgeschlossen.\n")
    print(log_message)

except Exception as e:
    import traceback
    print("\n❌ Error occurred / Fehler ist aufgetreten:")
    traceback.print_exc()

input("🔚 Press ENTER to exit / Drücke ENTER zum Beenden...")
