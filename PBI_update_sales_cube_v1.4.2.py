import pandas as pd
import os
import time
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime

start_time = time.time()

try:
    root = Tk()
    root.withdraw()

    # --- SELECT CSV FILE ---
    csv_path = filedialog.askopenfilename(
        title="Select CSV File / Wähle die CSV-Datei aus",
        filetypes=[("CSV files", "*.csv")]
    )
    if not csv_path:
        messagebox.showwarning("Cancelled / Abgebrochen", "❌ No CSV file selected.\n❌ Keine CSV-Datei ausgewählt.")
        raise SystemExit()

    # --- SELECT EXCEL FILE ---
    excel_path = filedialog.askopenfilename(
        title="Select Excel File / Wähle die Excel-Datei aus",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not excel_path:
        messagebox.showwarning("Cancelled / Abgebrochen", "❌ No Excel file selected.\n❌ Keine Excel-Datei ausgewählt.")
        raise SystemExit()

    #TODO: print the files being processed to the console
    #TODO: Print to the console the progress bar!

    # --- LOAD CSV DATA ---
    csv_df = pd.read_csv(csv_path)
    csv_headers = list(csv_df.columns)
    csv_data = csv_df.copy()

    # --- LOAD EXCEL WORKBOOK AND HEADERS ---
    wb = load_workbook(excel_path)
    sheet = wb.active

    # Find the first truly empty row based on column A
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=1).value is not None:
            start_row = row + 1
            break
    else:
        start_row = 2  # if the entire column A is empty, start at row 2

    excel_headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # --- Normalize Header Comparison ---
    def normalize_headers(header_row):
        return [str(h).strip().lower() for h in header_row]

    norm_csv_headers = normalize_headers(csv_headers)
    norm_excel_headers = normalize_headers(excel_headers)

    if norm_csv_headers == norm_excel_headers:
        #messagebox.showinfo(
        #    "Header Skipped / Kopfzeile übersprungen",
        #    "🧠 Header matched. The CSV header row will be skipped.\n🧠 Überschriften stimmen überein. Kopfzeile wird nicht angehängt."
        #)
        header_log = "🧠 Headers matched – CSV header row skipped.\n"
        csv_data = csv_df.iloc[1:].copy()

    else:
        mismatch_warning = (
            "⚠️ WARNING: Headers do not match!\n\n"
            "CSV Header:\n" + ", ".join(csv_headers) + "\n\n"
            "Excel Header:\n" + ", ".join(str(h) for h in excel_headers) + "\n\n"
            "⚠️ MISMATCH – Please verify before proceeding.\n"
            "⚠️ UNTERSCHIEDLICHE ÜBERSCHRIFTEN – Bitte überprüfen."
        )
        messagebox.showwarning("Header Mismatch / Überschriften stimmen nicht überein", mismatch_warning)
        header_log = "⚠️ Header mismatch – data appended anyway. Please review structure.\n"

    # --- Load Excel Data into DataFrame ---
    excel_data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    excel_df = pd.DataFrame(excel_data, columns=excel_headers)

    # --- Parse Dates & Filter Out Already Existing Days ---
    if 'Date' not in csv_data.columns or 'Date' not in excel_df.columns:
        raise ValueError("❌ 'Date' column not found in one of the files.\n❌ Spalte 'Date' fehlt in einer Datei.")

    csv_data['__date_only__'] = pd.to_datetime(csv_data['Date'], format='mixed', errors='coerce').dt.date
    excel_df['__date_only__'] = pd.to_datetime(excel_df['Date'], format='mixed', errors='coerce').dt.date

    existing_dates = set(excel_df['__date_only__'].dropna().unique())
    csv_data['__is_duplicate_date__'] = csv_data['__date_only__'].isin(existing_dates)

    skipped_dates = csv_data.loc[csv_data['__is_duplicate_date__'], '__date_only__'].unique()
    new_data = csv_data.loc[~csv_data['__is_duplicate_date__']].drop(columns=['__date_only__', '__is_duplicate_date__'])

    added_rows = len(new_data)
    skipped_dates_list = [str(d) for d in sorted(skipped_dates)]

    if added_rows == 0:
        messagebox.showinfo("No New Dates / Keine neuen Daten",
            "🔁 All dates in the CSV already exist in the Excel file.\n"
            "🔁 Alle Datumswerte aus der CSV-Datei sind bereits vorhanden.")
        raise SystemExit()

    # --- Write New Rows to Excel ---
    for row_idx, row in new_data.iterrows():
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=start_row + row_idx, column=col_idx, value=value)

    wb.save(excel_path)
    wb.close()

    # --- Logging ---
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elapsed = time.time() - start_time
    log_message = (
        f"{timestamp} | ✅ {added_rows} rows added to '{os.path.basename(excel_path)}'. "
        f"⏱️ Duration: {elapsed:.2f} seconds\n"
        f"{timestamp} | ✅ {added_rows} Zeilen zu '{os.path.basename(excel_path)}' hinzugefügt. "
        f"⏱️ Dauer: {elapsed:.2f} Sekunden\n"
        f"{header_log}"
    )

    if skipped_dates_list:
        skipped_str = ", ".join(skipped_dates_list)
        log_message += (
            f"❌ Skipped dates (already in Excel): {skipped_str}\n"
            f"❌ Übersprungene Datumswerte (bereits vorhanden): {skipped_str}\n"
        )
    log_path = os.path.join(os.path.dirname(__file__), "import_log.txt")
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(log_message)

    # --- Confirmation Popup ---
    popup_message = (
        # English block
        f"✅ {added_rows} rows added.\n"
        f"📄 File: {os.path.basename(csv_path)}\n"
        f"📊 Target: {os.path.basename(excel_path)}\n"
        f"📈 Sheet starts at row {start_row}\n"
        f"⏱️ Duration: {elapsed:.2f} seconds\n\n"
        
        # German block
        f"✅ {added_rows} Zeilen hinzugefügt.\n"
        f"📄 Datei: {os.path.basename(csv_path)}\n"
        f"📊 Ziel: {os.path.basename(excel_path)}\n"
        f"📈 Arbeitsblatt beginnt bei Zeile {start_row}\n"
        f"⏱️ Dauer: {elapsed:.2f} Sekunden\n\n"
        
        f"{header_log.strip()}"
    )

    if skipped_dates_list:
        popup_message += (
            f"\n❌ Skipped dates:\n" +
            "\n".join(skipped_dates_list)
        )

    popup_message += "\n\n(Log entry created / Protokolleintrag erstellt)"
    messagebox.showinfo("Import Complete / Import abgeschlossen", popup_message)

except Exception as e:
    import traceback
    print("\n❌ Error occurred / Fehler ist aufgetreten:")
    traceback.print_exc()

    input("🔚 Press ENTER to exit / Drücke ENTER zum Beenden...")
