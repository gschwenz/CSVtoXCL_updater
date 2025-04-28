import os
import pandas as pd
from openpyxl import load_workbook

# --- USER CONFIG ---
folder_path = r"C:\Users\Artkaemper\SPIGGLE-THEIS MEDIZINTECHNIK GMBH\620 USA + Bryan Medical Integration - 06 - Administrative and Operations - 06 - Administrative and Operations\Power BI\03_Daily Files"
sales_path = r"C:\Users\Artkaemper\SPIGGLE-THEIS MEDIZINTECHNIK GMBH\620 USA + Bryan Medical Integration - 06 - Administrative and Operations - 06 - Administrative and Operations\Power BI\02_Master Sales File\Sales_Cube_BM_Master.xlsx"

# --- FIND FIRST CSV FILE ---
csv_file = next((f for f in os.listdir(folder_path) if f.lower().endswith('.csv')), None)
if not csv_file:
    raise FileNotFoundError("Keine CSV-Datei im Ordner gefunden!")

csv_path = os.path.join(folder_path, csv_file)
print(f"Gefundene Datei: {csv_path}")

# --- LOAD DATA ---
csv_df = pd.read_csv(csv_path)

# Load existing Excel workbook and get the first sheet as a DataFrame
book = load_workbook(sales_path)
writer = pd.ExcelWriter(sales_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

sheet_name = book.sheetnames[0]
existing_df = pd.read_excel(sales_path, sheet_name=sheet_name)

# --- APPEND CSV TO EXISTING DATA ---
updated_df = pd.concat([existing_df, csv_df], ignore_index=True)

# --- WRITE BACK TO EXCEL ---
updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
writer.close()

print("Daten erfolgreich angehängt!\nData attached successfully!")
input("Press Enter to exit...")
