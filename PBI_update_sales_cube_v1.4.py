from openpyxl import load_workbook
import pandas as pd
from tkinter import filedialog, messagebox, Tk
from datetime import datetime
import os

import time
start_time = time.time()


try:
    root = Tk()
    root.withdraw()

    csv_path = filedialog.askopenfilename(title="Wähle CSV-Datei", filetypes=[("CSV files", "*.csv")])
    if not csv_path:
        messagebox.showwarning("Abgebrochen", "❌ Keine CSV-Datei ausgewählt.")
        raise SystemExit()

    excel_path = filedialog.askopenfilename(title="Wähle Excel-Datei", filetypes=[("Excel files", "*.xlsx")])
    if not excel_path:
        messagebox.showwarning("Abgebrochen", "❌ Keine Excel-Datei ausgewählt.")
        raise SystemExit()

    csv_df = pd.read_csv(csv_path)
    added_rows = len(csv_df)

    wb = load_workbook(excel_path)
    sheet = wb.active

    # Find the first empty row in Excel sheet
    start_row = sheet.max_row + 1

    # Write data directly to the Excel worksheet
    for row_idx, row in csv_df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=start_row + row_idx, column=col_idx, value=value)

    wb.save(excel_path)
    wb.close()

    # Log and show message
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_message = (
        f"{timestamp} | ✅ {added_rows} Zeilen von '{os.path.basename(csv_path)}' "
        f"zu '{os.path.basename(excel_path)}' (ab Zeile {start_row}) hinzugefügt.\n"
    )
    log_path = os.path.join(os.path.dirname(__file__), "import_log.txt")
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(log_message)

    messagebox.showinfo("Import abgeschlossen", log_message)

except Exception as e:
    import traceback
    print("\n❌ Ein Fehler ist aufgetreten:")
    traceback.print_exc()
    input("🔚 Drücke ENTER zum Beenden...")

elapsed = time.time() - start_time
print(f"\n⏱️ Verarbeitungszeit: {elapsed:.2f} Sekunden")
