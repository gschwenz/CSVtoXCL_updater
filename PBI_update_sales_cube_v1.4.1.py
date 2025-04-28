import pandas as pd
import os
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime
import time

start_time = time.time()

try:
    root = Tk()
    root.withdraw()

    # --- SELECT CSV FILE ---
    csv_path = filedialog.askopenfilename(
        title="Select CSV File / Wähle die CSV-Datei aus",
        filetypes=[("CSV files", "*.csv")]
    )
    if not csv_path:
        messagebox.showwarning("Cancelled / Abgebrochen", "❌ No CSV file selected.\n❌ Keine CSV-Datei ausgewählt.")
        raise SystemExit()

    # --- SELECT EXCEL FILE ---
    excel_path = filedialog.askopenfilename(
        title="Select Excel File / Wähle die Excel-Datei aus",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not excel_path:
        messagebox.showwarning("Cancelled / Abgebrochen", "❌ No Excel file selected.\n❌ Keine Excel-Datei ausgewählt.")
        raise SystemExit()

    # --- LOAD CSV ---
    csv_df = pd.read_csv(csv_path)
    csv_headers = list(csv_df.columns)
    csv_data = csv_df.copy()

    # --- LOAD EXCEL & HEADER ---
    wb = load_workbook(excel_path)
    sheet = wb.active
    start_row = sheet.max_row + 1

    excel_headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]


    # --- HEADER COMPARISON LOGIC ---
    header_log = ""
    def normalize_headers(header_row):
        return [str(h).strip().lower() for h in header_row]

    norm_csv_headers = normalize_headers(csv_headers)
    norm_excel_headers = normalize_headers(excel_headers)

    if norm_csv_headers == norm_excel_headers:
        #messagebox.showinfo(
        #    "Header Skipped / Kopfzeile übersprungen",
        #    "🧠 Header matched. The CSV header row will be skipped.\n🧠 Überschriften stimmen überein. Kopfzeile wird nicht angehängt."
        #)
        header_log = "🧠 Headers matched – CSV header row skipped.\n"
        csv_data = csv_df.iloc[1:]
    else:
        mismatch_warning = (
            "⚠️ WARNING: Headers do not match!\n\n"
            "CSV Header:\n" + ", ".join(csv_headers) + "\n\n"
            "Excel Header:\n" + ", ".join(str(h) for h in excel_headers) + "\n\n"
            "⚠️ MISMATCH – Please verify before proceeding.\n"
            "⚠️ UNTERSCHIEDLICHE ÜBERSCHRIFTEN – Bitte überprüfen."
        )
        messagebox.showwarning("Header Mismatch / Überschriften stimmen nicht überein", mismatch_warning)
        header_log = "⚠️ Header mismatch – data appended anyway. Please review structure.\n"


    added_rows = len(csv_data)

    # --- WRITE TO EXCEL ---
    for row_idx, row in csv_data.iterrows():
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=start_row + row_idx, column=col_idx, value=value)

    wb.save(excel_path)
    wb.close()

    # --- LOGGING ---
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elapsed = time.time() - start_time
    log_message = (
        f"{timestamp} | ✅ {added_rows} rows from '{os.path.basename(csv_path)}' "
        f"added to '{os.path.basename(excel_path)}' (starting at row {start_row}). "
        f"⏱️ Duration: {elapsed:.2f} seconds\n"
        f"{timestamp} | ✅ {added_rows} Zeilen von '{os.path.basename(csv_path)}' "
        f"zu '{os.path.basename(excel_path)}' (ab Zeile {start_row}) hinzugefügt. "
        f"⏱️ Dauer: {elapsed:.2f} Sekunden\n"
        f"{header_log}\n"
    )

    log_path = os.path.join(os.path.dirname(__file__), "import_log.txt")
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(log_message)

    # --- FINAL POPUP ---
    messagebox.showinfo(
        "Import Complete / Import abgeschlossen",
        f"✅ {added_rows} rows added.\n✅ {added_rows} Zeilen hinzugefügt.\n\n"
        f"📄 File: {os.path.basename(csv_path)}\n📄 Datei: {os.path.basename(csv_path)}\n"
        f"📊 Target: {os.path.basename(excel_path)}\n📊 Ziel: {os.path.basename(excel_path)}\n"
        f"📈 Sheet starts at row {start_row}\n📈 Arbeitsblatt beginnt bei Zeile {start_row}\n"
        f"⏱️ Duration: {elapsed:.2f} seconds\n⏱️ Dauer: {elapsed:.2f} Sekunden\n\n"
        f"{header_log.strip()}\n"
        f"(Log entry created / Protokolleintrag erstellt)"
    )

except Exception as e:
    import traceback
    print("\n❌ Error occurred / Fehler ist aufgetreten:")
    traceback.print_exc()

    input("🔚 Press ENTER to exit / Drücke ENTER zum Beenden...")
